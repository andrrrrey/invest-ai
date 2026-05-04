import csv
import io
from datetime import datetime
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from pydantic import BaseModel
from sqlalchemy.orm import Session

from ...database import get_db
from ...models.fact_entry import FactEntry
from ...models.project import Project
from ...auth import get_current_user
from ...services import ai_service
from ... import settings_store

router = APIRouter(tags=["fact"])


# ── Schemas ───────────────────────────────────────────────────────────────────

class FactEntryIn(BaseModel):
    year: int
    month: int
    metric_name: str
    plan_value: Optional[float] = None
    fact_value: Optional[float] = None


class FactEntryRead(BaseModel):
    id: int
    year: int
    month: int
    metric_name: str
    plan_value: Optional[float]
    fact_value: Optional[float]
    deviation_abs: Optional[float]   # fact - plan
    deviation_pct: Optional[float]   # (fact - plan) / plan * 100
    updated_at: Optional[datetime]

    class Config:
        from_attributes = True


# ── Helpers ───────────────────────────────────────────────────────────────────

def _deviation(plan: float | None, fact: float | None) -> tuple[float | None, float | None]:
    if plan is None or fact is None:
        return None, None
    diff = fact - plan
    pct = (diff / plan * 100) if plan != 0 else None
    return round(diff, 4), round(pct, 2) if pct is not None else None


def _to_read(e: FactEntry) -> FactEntryRead:
    dev_abs, dev_pct = _deviation(e.plan_value, e.fact_value)
    return FactEntryRead(
        id=e.id,
        year=e.year,
        month=e.month,
        metric_name=e.metric_name,
        plan_value=e.plan_value,
        fact_value=e.fact_value,
        deviation_abs=dev_abs,
        deviation_pct=dev_pct,
        updated_at=e.updated_at,
    )


def _upsert(db: Session, project_id: int, item: FactEntryIn) -> FactEntry:
    existing = (
        db.query(FactEntry)
        .filter_by(project_id=project_id, year=item.year, month=item.month, metric_name=item.metric_name)
        .first()
    )
    if existing:
        if item.plan_value is not None:
            existing.plan_value = item.plan_value
        if item.fact_value is not None:
            existing.fact_value = item.fact_value
        return existing
    e = FactEntry(
        project_id=project_id,
        year=item.year,
        month=item.month,
        metric_name=item.metric_name,
        plan_value=item.plan_value,
        fact_value=item.fact_value,
    )
    db.add(e)
    return e


# ── Endpoints ─────────────────────────────────────────────────────────────────

@router.get("/projects/{project_id}/fact", response_model=List[FactEntryRead])
def get_fact(
    project_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    if not db.get(Project, project_id):
        raise HTTPException(status_code=404, detail="Проект не найден")
    rows = (
        db.query(FactEntry)
        .filter(FactEntry.project_id == project_id)
        .order_by(FactEntry.metric_name, FactEntry.year, FactEntry.month)
        .all()
    )
    return [_to_read(e) for e in rows]


@router.put("/projects/{project_id}/fact", response_model=List[FactEntryRead])
def upsert_fact(
    project_id: int,
    items: List[FactEntryIn],
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    if not db.get(Project, project_id):
        raise HTTPException(status_code=404, detail="Проект не найден")
    updated = [_upsert(db, project_id, item) for item in items]
    db.commit()
    for e in updated:
        db.refresh(e)
    return [_to_read(e) for e in updated]


@router.post("/projects/{project_id}/fact/import", response_model=List[FactEntryRead])
async def import_fact(
    project_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    if not db.get(Project, project_id):
        raise HTTPException(status_code=404, detail="Проект не найден")

    content = await file.read()
    filename = (file.filename or "").lower()

    items: List[FactEntryIn] = []

    if filename.endswith(".xlsx") or filename.endswith(".xls"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
            headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = dict(zip(headers, row))
                try:
                    items.append(FactEntryIn(
                        year=int(row_dict.get("year") or row_dict.get("год", 0)),
                        month=int(row_dict.get("month") or row_dict.get("месяц", 0)),
                        metric_name=str(row_dict.get("metric_name") or row_dict.get("метрика", "")),
                        fact_value=float(row_dict.get("value") or row_dict.get("значение") or 0),
                    ))
                except (TypeError, ValueError):
                    continue
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Ошибка разбора Excel: {e}")
    else:
        # CSV — expected columns: year, month, metric_name, value
        try:
            text = content.decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(text))
            for row in reader:
                try:
                    items.append(FactEntryIn(
                        year=int(row.get("year") or row.get("год", 0)),
                        month=int(row.get("month") or row.get("месяц", 0)),
                        metric_name=str(row.get("metric_name") or row.get("метрика", "")),
                        fact_value=float(row.get("value") or row.get("значение") or 0),
                    ))
                except (TypeError, ValueError):
                    continue
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Ошибка разбора CSV: {e}")

    if not items:
        raise HTTPException(status_code=400, detail="Файл не содержит корректных данных")

    updated = [_upsert(db, project_id, item) for item in items]
    db.commit()
    for e in updated:
        db.refresh(e)
    return [_to_read(e) for e in updated]


@router.get("/projects/{project_id}/fact/forecast", response_model=List[FactEntryRead])
def get_forecast(
    project_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """
    Simple linear trend forecast: for each metric with at least 2 fact points,
    compute average month-over-month delta and project forward to fill missing months.
    Returns only the projected (non-existing) entries as a preview — does not write to DB.
    """
    if not db.get(Project, project_id):
        raise HTTPException(status_code=404, detail="Проект не найден")

    rows = (
        db.query(FactEntry)
        .filter(FactEntry.project_id == project_id, FactEntry.fact_value.isnot(None))
        .order_by(FactEntry.metric_name, FactEntry.year, FactEntry.month)
        .all()
    )

    # Group by metric
    from collections import defaultdict
    by_metric: dict[str, list[FactEntry]] = defaultdict(list)
    for r in rows:
        by_metric[r.metric_name].append(r)

    projections: list[FactEntryRead] = []
    for metric, entries in by_metric.items():
        if len(entries) < 2:
            continue
        # Average delta
        deltas = [entries[i + 1].fact_value - entries[i].fact_value for i in range(len(entries) - 1)]
        avg_delta = sum(deltas) / len(deltas)
        last = entries[-1]
        last_year, last_month = last.year, last.month
        last_val = last.fact_value

        # Project 6 months forward
        for i in range(1, 7):
            m = last_month + i
            y = last_year + (m - 1) // 12
            m = ((m - 1) % 12) + 1
            proj_val = round(last_val + avg_delta * i, 4)
            projections.append(FactEntryRead(
                id=0,
                year=y,
                month=m,
                metric_name=metric,
                plan_value=None,
                fact_value=proj_val,
                deviation_abs=None,
                deviation_pct=None,
                updated_at=None,
            ))

    return projections


@router.post("/projects/{project_id}/fact/ai-commentary")
def ai_commentary(
    project_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
) -> dict:
    provider = settings_store.get_ai_provider()
    key = settings_store.get_anthropic_key() if provider == "anthropic" else settings_store.get_openai_key()
    if not key:
        raise HTTPException(status_code=503, detail="AI не настроен")

    project = db.get(Project, project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Проект не найден")

    rows = (
        db.query(FactEntry)
        .filter(FactEntry.project_id == project_id)
        .order_by(FactEntry.metric_name, FactEntry.year, FactEntry.month)
        .all()
    )

    fact_summary = "\n".join(
        f"  {e.metric_name} {e.year}/{e.month:02d}: план={e.plan_value}, факт={e.fact_value}, "
        f"откл={round(e.fact_value - e.plan_value, 2) if e.plan_value and e.fact_value else 'н/д'}"
        for e in rows
        if e.fact_value is not None
    ) or "Фактических данных нет."

    try:
        result = ai_service.analyze_project(
            project={
                "name": project.name,
                "status": project.status,
                "project_type": project.project_type,
                "fact_summary": fact_summary,
            },
            metrics=project.metrics or {},
        )
        return result
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Ошибка AI: {e}")
