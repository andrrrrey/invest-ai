"""
Excel export service — builds a .xlsx workbook for portfolio management reports.
"""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PURPLE_HEX = "FF5E5CE6"
LIGHT_GRAY_HEX = "FFF2F2F2"
RED_HEX = "FFFFD7D7"
YELLOW_HEX = "FFFFF3CD"
GREEN_HEX = "FFD4EDDA"
WHITE_HEX = "FFFFFFFF"

_THIN = Side(style="thin", color="FFD0D0D0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _header_fill() -> PatternFill:
    return PatternFill("solid", fgColor=PURPLE_HEX)


def _gray_fill() -> PatternFill:
    return PatternFill("solid", fgColor=LIGHT_GRAY_HEX)


def _apply_header(ws, values: list[str], row: int = 1):
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = Font(bold=True, color=WHITE_HEX)
        cell.fill = _header_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER


def _auto_width(ws, min_width: int = 10, max_width: int = 40):
    for col in ws.columns:
        max_len = min_width
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), max_width))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max_len + 2


def _risk_fill(risk_level: str) -> PatternFill | None:
    rl = (risk_level or "").lower()
    if "высок" in rl:
        return PatternFill("solid", fgColor=RED_HEX)
    if "средн" in rl or "умерен" in rl:
        return PatternFill("solid", fgColor=YELLOW_HEX)
    if "низк" in rl:
        return PatternFill("solid", fgColor=GREEN_HEX)
    return None


def _get_risk_level(project: dict) -> str:
    risks = project.get("risks_data") or {}
    ai = risks.get("ai_assessment") or {}
    return ai.get("risk_level") or risks.get("overall_risk") or "—"


_ROUTE_LABELS = {
    "fast_track": "FAST TRACK",
    "efficiency_play": "EFFICIENCY PLAY",
    "backlog_stop": "BACKLOG / STOP",
    "manual_review": "Ручная проверка",
}


def _value_drivers_text(fm: dict) -> str:
    drivers = fm.get("op_value_drivers") or []
    other = fm.get("op_value_drivers_other") or ""
    out = [other if d == "__other__" else d for d in drivers]
    return "; ".join(str(d) for d in out if d)


def _has_type(projects: list[dict], ptype: str) -> bool:
    return any(p.get("project_type") == ptype for p in projects)


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _sheet_portfolio_summary(wb: Workbook, stats: dict, period_label: str):
    ws = wb.active
    ws.title = "Сводка портфеля"

    # Title row
    ws.merge_cells("A1:B1")
    title_cell = ws["A1"]
    title_cell.value = f"Портфель: {period_label}"
    title_cell.font = Font(bold=True, color=WHITE_HEX, size=13)
    title_cell.fill = _header_fill()
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    by_status = stats.get("by_status", {})
    by_type   = stats.get("by_type", {})

    rows = [
        ("Всего проектов", stats.get("total", 0)),
        ("", ""),
        ("По статусу", ""),
        ("  Черновик",              by_status.get("draft", 0)),
        ("  На рассмотрении",       by_status.get("pending_approval", 0)),
        ("  Одобрено",              by_status.get("approved", 0)),
        ("  Отклонено",             by_status.get("rejected", 0)),
        ("", ""),
        ("По типу", ""),
        ("  Инвестиционные",        by_type.get("investment", 0)),
        ("  Операционные",          by_type.get("operational", 0)),
        ("", ""),
        ("Суммарный NPV (₽)",       stats.get("total_npv", 0)),
        ("Средний IRR (%)",         stats.get("avg_irr", "н/д")),
        ("Высокорисковых проектов", stats.get("high_risk_count", 0)),
        ("", ""),
        ("Инвестиционный бюджет (₽)",   stats.get("investment_budget", "не задан")),
        ("Одобренные инвестиции (₽)",   stats.get("approved_investment", 0)),
        ("Доступно для инвестиций (₽)", stats.get("available_for_investment", "н/д")),
    ]

    for i, (label, value) in enumerate(rows, start=2):
        label_cell = ws.cell(row=i, column=1, value=label)
        value_cell = ws.cell(row=i, column=2, value=value)
        if label and not label.startswith(" ") and label not in ("По статусу", "По типу"):
            label_cell.font = Font(bold=True)
        if label in ("По статусу", "По типу"):
            label_cell.font = Font(bold=True, italic=True)
            label_cell.fill = _gray_fill()
            value_cell.fill = _gray_fill()
        label_cell.border = _BORDER
        value_cell.border = _BORDER
        value_cell.alignment = Alignment(horizontal="right")

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 22


def _sheet_all_projects(wb: Workbook, projects: list[dict]):
    ws = wb.create_sheet("Все проекты")

    headers = [
        # Common
        "ID", "Название", "Бизнес-юнит", "Тип", "Стадия", "Статус", "Владелец",
        "Дата начала", "Платформа",
        # Investment metrics
        "NPV (₽)", "IRR (%)", "DPP (лет)", "PI", "LTV/CAC", "CAC (₽)", "ARPA (₽)",
        "LTV (₽)", "Отток (%)", "Gross Margin (%)", "Initial Investment (₽)",
        # Operational application
        "Категория (опер.)", "Тип инвестиции", "МВЗ", "Ключевая метрика",
        "Baseline", "Target", "Value Drivers", "Бюджет заявки (₽)",
        # Value Score
        "Value Score", "Зона", "EV (×3)", "Срочность (×2)", "ROI (×1)",
        "Масштаб. (×2)", "Unit-эк. (×2)",
        # Risk / routing
        "Уровень риска", "Маршрут решения",
        # Smart contract
        "SC: тип", "SC: структура", "SC: вознаграждение (₽)", "SC: coins",
    ]
    _apply_header(ws, headers)
    ws.freeze_panes = "A2"

    for p in projects:
        m = p.get("metrics") or {}
        fm = p.get("financial_model") or {}
        vs = p.get("value_score_data") or {}
        sc = p.get("smart_contract_data") or {}
        risk_level = _get_risk_level(p)
        mvz = " / ".join(
            str(fm.get(k)) for k in ("op_mvz_main", "op_mvz_sub1", "op_mvz_sub2") if fm.get(k)
        )
        budget_total = 0.0
        for brow in (fm.get("op_budget_rows") or []):
            for v in (brow.get("annual_totals") or []):
                try:
                    budget_total += float(v or 0)
                except (TypeError, ValueError):
                    pass

        row = [
            p.get("id"), p.get("name"), p.get("business_unit"), p.get("project_type"),
            p.get("stage"), p.get("status"), p.get("owner"),
            str(p.get("start_date") or ""), p.get("platform"),
            m.get("npv"), m.get("irr"), m.get("dpp"), m.get("pi"), m.get("ltvCac"),
            m.get("cac"), m.get("arpa"), m.get("ltv"), m.get("avgChurn"),
            m.get("grossMargin"), fm.get("initialInvestment"),
            fm.get("op_category"), fm.get("op_investment_type"), mvz, fm.get("op_metrics"),
            fm.get("op_baseline"), fm.get("op_target"), _value_drivers_text(fm),
            (budget_total or None),
            vs.get("total"), vs.get("band"), vs.get("ev_impact"), vs.get("urgency"),
            vs.get("roi_predictability"), vs.get("scalability"), vs.get("unit_economics"),
            risk_level, _ROUTE_LABELS.get(p.get("decision_route"), p.get("decision_route")),
            sc.get("scType"), sc.get("contractType"),
            sc.get("totalRewardRub"), sc.get("totalCoins"),
        ]
        ws.append(row)
        row_idx = ws.max_row
        fill = _risk_fill(risk_level)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = _BORDER
            if fill:
                cell.fill = fill

    _auto_width(ws, max_width=30)


def _sheet_annual_cash_flows(wb: Workbook, projects: list[dict]):
    ws = wb.create_sheet("Денежные потоки")

    year_headers = [f"Год {i}" for i in range(6)]
    top_headers = ["Проект", "Показатель"] + year_headers
    _apply_header(ws, top_headers)
    ws.freeze_panes = "A2"

    metric_keys = [
        ("annualRevenue",       "Выручка (₽)"),
        ("totalCosts",          "Затраты (₽)"),
        ("netCashFlow",         "Чистый ДП (₽)"),
        ("discountedCF",        "Дисконт. ДП (₽)"),
        ("cumulativeDcfSeries", "Накопл. ДДП (₽)"),
    ]

    for p in projects:
        if p.get("project_type") == "operational":
            continue
        m = p.get("metrics") or {}
        name = p.get("name", "—")
        first_row = True
        for key, label in metric_keys:
            series = m.get(key) or []
            # Pad or trim to 6 values (Year 0..5)
            values = list(series[:6]) + [None] * max(0, 6 - len(series))
            row_data = [name if first_row else "", label] + values
            ws.append(row_data)
            row_idx = ws.max_row
            for col_idx in range(1, len(top_headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = _BORDER
                if first_row:
                    cell.fill = _gray_fill()
            first_row = False

    _auto_width(ws)


def _sheet_quarterly_user_metrics(wb: Workbook, projects: list[dict]):
    ws = wb.create_sheet("Квартальные метрики")

    quarter_headers = [f"Y{y+1}Q{q+1}" for y in range(5) for q in range(4)]
    top_headers = ["Проект", "Показатель"] + quarter_headers
    _apply_header(ws, top_headers)
    ws.freeze_panes = "A2"

    for p in projects:
        name = p.get("name", "—")
        m = p.get("metrics") or {}

        if p.get("project_type") == "operational":
            ws.append([name, "Операционный проект — квартальные данные недоступны"])
            ws.cell(ws.max_row, 1).font = Font(italic=True, color="FF888888")
            continue

        paid_users = m.get("paidUsers") or []
        revenue    = m.get("revenue") or []

        def flatten_matrix(matrix):
            result = []
            for year_data in matrix[:5]:
                if isinstance(year_data, list):
                    result.extend(year_data[:4])
                    result.extend([None] * max(0, 4 - len(year_data)))
                else:
                    result.extend([None, None, None, None])
            return result + [None] * max(0, 20 - len(result))

        pu_flat  = flatten_matrix(paid_users)
        rev_flat = flatten_matrix(revenue)

        for label, values in [("Платные пользователи", pu_flat), ("Выручка (₽)", rev_flat)]:
            ws.append([name, label] + values[:20])
            row_idx = ws.max_row
            for col_idx in range(1, len(top_headers) + 1):
                ws.cell(row=row_idx, column=col_idx).border = _BORDER
            name = ""  # only print name once

    _auto_width(ws, max_width=15)


def _sheet_operational_applications(wb: Workbook, projects: list[dict]):
    """One row per operational application (заявка) with all op_* / Value Score / Risk fields."""
    ws = wb.create_sheet("Заявки (операционные)")
    headers = [
        "ID", "Название", "Категория", "МВЗ (осн)", "МВЗ (суб1)", "МВЗ (суб2)",
        "Тип инвестиции", "Запрашиваемый ресурс", "Инвестиционный тезис",
        "Value Drivers", "Ключевая метрика", "Baseline", "Target",
        "Экономика / Payback", "Stop-loss", "Stage-gates", "Бюджет заявки (₽)",
        "Value Score", "Зона", "EV (×3)", "Срочность (×2)", "ROI (×1)",
        "Масштаб. (×2)", "Unit-эк. (×2)",
        "Общий риск", "Риск исполнения", "Рыночный", "Финансовый", "Стратегический",
        "Маршрут", "Комментарий AI",
    ]
    _apply_header(ws, headers)
    ws.freeze_panes = "A2"

    for p in projects:
        if p.get("project_type") != "operational":
            continue
        fm = p.get("financial_model") or {}
        vs = p.get("value_score_data") or {}
        risks = p.get("risks_data") or {}
        rs = risks.get("risk_scores") or {}
        budget_total = 0.0
        for brow in (fm.get("op_budget_rows") or []):
            for v in (brow.get("annual_totals") or []):
                try:
                    budget_total += float(v or 0)
                except (TypeError, ValueError):
                    pass
        ws.append([
            p.get("id"), p.get("name"), fm.get("op_category"),
            fm.get("op_mvz_main"), fm.get("op_mvz_sub1"), fm.get("op_mvz_sub2"),
            fm.get("op_investment_type"), fm.get("op_requested_resource"),
            fm.get("op_investment_thesis"), _value_drivers_text(fm),
            fm.get("op_metrics"), fm.get("op_baseline"), fm.get("op_target"),
            fm.get("op_economics"), fm.get("op_stop_loss"), fm.get("op_stage_gates"),
            (budget_total or None),
            vs.get("total"), vs.get("band"), vs.get("ev_impact"), vs.get("urgency"),
            vs.get("roi_predictability"), vs.get("scalability"), vs.get("unit_economics"),
            risks.get("overall_risk"), rs.get("execution_risk"), rs.get("market_risk"),
            rs.get("financial_risk"), rs.get("strategic_risk"),
            _ROUTE_LABELS.get(p.get("decision_route"), p.get("decision_route")),
            risks.get("commentary"),
        ])
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=ws.max_row, column=col_idx).border = _BORDER

    _auto_width(ws, max_width=45)


def _sheet_application_budgets(wb: Workbook, projects: list[dict]):
    """Annual budget breakdown for operational applications (статья × годы)."""
    rows_exist = any(
        (p.get("financial_model") or {}).get("op_budget_rows")
        for p in projects if p.get("project_type") == "operational"
    )
    if not rows_exist:
        return
    ws = wb.create_sheet("Бюджет заявок")
    max_years = 1
    for p in projects:
        for brow in ((p.get("financial_model") or {}).get("op_budget_rows") or []):
            max_years = max(max_years, len(brow.get("annual_totals") or []))
    headers = ["Заявка", "Статья"] + [f"Год {i+1} (₽)" for i in range(max_years)] + ["Итого (₽)"]
    _apply_header(ws, headers)
    ws.freeze_panes = "A2"

    for p in projects:
        if p.get("project_type") != "operational":
            continue
        name = p.get("name", "—")
        first = True
        for brow in ((p.get("financial_model") or {}).get("op_budget_rows") or []):
            totals = list(brow.get("annual_totals") or [])
            padded = totals + [None] * max(0, max_years - len(totals))
            row_total = sum(float(v or 0) for v in totals) if totals else None
            ws.append([name if first else "", brow.get("category")] + padded[:max_years] + [row_total])
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=col_idx).border = _BORDER
            first = False

    _auto_width(ws, max_width=22)


def _sheet_smart_contracts(wb: Workbook, projects: list[dict]):
    ws = wb.create_sheet("Смарт-контракты")
    headers = [
        "ID", "Название", "Тип (scType)", "Структура", "Куратор", "Статус",
        "Краткое описание", "Бизнес-эффект", "Набор команды", "Курс NAV",
        "Итого вознаграждение (₽)", "Итого coins",
        "Фикс. оклад/мес (₽)", "Итого оклад (₽)", "Итого премии (₽)", "Grand total (₽)",
    ]
    _apply_header(ws, headers)
    ws.freeze_panes = "A2"

    for p in projects:
        if p.get("project_type") != "smart_contract":
            continue
        sc = p.get("smart_contract_data") or {}
        lt = sc.get("leadershipTotals") or {}
        ws.append([
            p.get("id"), sc.get("name") or p.get("name"), sc.get("scType"),
            sc.get("contractType"), sc.get("curator"), p.get("status"),
            sc.get("shortDescription"), sc.get("businessEffect"),
            "Да" if sc.get("needsRecruitment") else "Нет", sc.get("navRate"),
            sc.get("totalRewardRub"), sc.get("totalCoins"),
            sc.get("fixedSalaryMonthly"), lt.get("totalSalary"),
            lt.get("totalBonus"), lt.get("grand"),
        ])
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=ws.max_row, column=col_idx).border = _BORDER

    _auto_width(ws, max_width=45)


def _sheet_smart_contract_milestones(wb: Workbook, projects: list[dict]):
    rows_exist = any(
        (p.get("smart_contract_data") or {}).get("milestones")
        for p in projects if p.get("project_type") == "smart_contract"
    )
    if not rows_exist:
        return
    ws = wb.create_sheet("Milestones")
    headers = [
        "Контракт", "Этап", "Описание", "Дедлайн", "Вознаграждение (₽)", "Coins", "SMART",
    ]
    _apply_header(ws, headers)
    ws.freeze_panes = "A2"

    for p in projects:
        if p.get("project_type") != "smart_contract":
            continue
        sc = p.get("smart_contract_data") or {}
        cname = sc.get("name") or p.get("name", "—")
        first = True
        for ms in (sc.get("milestones") or []):
            crit = ms.get("criteria") or {}
            smart = ", ".join(k for k, v in crit.items() if v) if isinstance(crit, dict) else ""
            ws.append([
                cname if first else "", ms.get("title"), ms.get("description"),
                ms.get("targetDate"), ms.get("rewardRub"), ms.get("coins"), smart,
            ])
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=col_idx).border = _BORDER
            first = False

    _auto_width(ws, max_width=40)


def _sheet_smart_contract_team(wb: Workbook, projects: list[dict]):
    rows_exist = any(
        (p.get("smart_contract_data") or {}).get("team")
        for p in projects if p.get("project_type") == "smart_contract"
    )
    if not rows_exist:
        return
    ws = wb.create_sheet("Команда и Coins")
    headers = ["Контракт", "Участник", "Роль", "Coins"]
    _apply_header(ws, headers)
    ws.freeze_panes = "A2"

    for p in projects:
        if p.get("project_type") != "smart_contract":
            continue
        sc = p.get("smart_contract_data") or {}
        cname = sc.get("name") or p.get("name", "—")
        first = True
        for t in (sc.get("team") or []):
            ws.append([cname if first else "", t.get("name"), t.get("role"), t.get("coins")])
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=col_idx).border = _BORDER
            first = False

    _auto_width(ws, max_width=30)


def _sheet_ai_commentary(wb: Workbook, projects: list[dict], ai_commentaries: dict):
    ws = wb.create_sheet("AI-комментарии")
    _apply_header(ws, ["Проект", "AI-комментарий"])
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 80

    for p in projects:
        commentary = ai_commentaries.get(p.get("id"), "—")
        ws.append([p.get("name", "—"), commentary])
        row_idx = ws.max_row
        ws.row_dimensions[row_idx].height = 60
        for col_idx in (1, 2):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = _BORDER


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def build_excel(
    projects: list[dict],
    stats: dict,
    detail_level: str,
    include_ai: bool,
    ai_commentaries: dict,
    period_label: str,
) -> io.BytesIO:
    """
    Build a .xlsx workbook and return it as a BytesIO stream.

    :param projects:        List of project dicts (id, name, metrics, risks_data, …)
    :param stats:           Portfolio stats dict (same shape as GET /api/v1/stats/)
    :param detail_level:    "summary" or "full"
    :param include_ai:      Whether to add the AI commentary sheet
    :param ai_commentaries: {project_id: commentary_text}
    :param period_label:    Human-readable period string, e.g. "Q1 2026"
    """
    wb = Workbook()

    _sheet_portfolio_summary(wb, stats, period_label)
    _sheet_all_projects(wb, projects)

    has_operational = _has_type(projects, "operational")
    has_investment  = _has_type(projects, "investment")
    has_smart       = _has_type(projects, "smart_contract")

    # Operational application detail (заявки) — always when present
    if has_operational:
        _sheet_operational_applications(wb, projects)
        _sheet_application_budgets(wb, projects)

    # Smart contract detail — always when present
    if has_smart:
        _sheet_smart_contracts(wb, projects)
        _sheet_smart_contract_milestones(wb, projects)
        _sheet_smart_contract_team(wb, projects)

    if detail_level == "full" and has_investment:
        _sheet_annual_cash_flows(wb, projects)
        _sheet_quarterly_user_metrics(wb, projects)

    if include_ai:
        _sheet_ai_commentary(wb, projects, ai_commentaries)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
